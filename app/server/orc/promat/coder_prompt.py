from app.server.orc.promat.nn_promat_core import FULL_NN_PROMAT


# ──────────────────────────────────────────────────────────────────────────────
#  Ready-to-use Python helpers implementing the 7-layer NN logic.
#  The Coder Agent generates code using these patterns.
# ──────────────────────────────────────────────────────────────────────────────

_NN_PYTHON_HELPERS = '''
# ── Layer 0+1: structural extraction ─────────────────────────────────────────
import math
import re
from collections import Counter
from openpyxl import load_workbook

REQUIRED_SECTIONS = frozenset([
    "assets", "current assets", "long-term assets",
    "liabilities and equity", "current liabilities",
    "long-term liabilities", "equity",
])
COMPANY_KW = frozenset(["nis", "dollar", "$", "usd", "ils", "inc", "ltd"])
AJE_KW = frozenset(["aje", "adjusting", "adjustment"])
CONSOL_KW = frozenset(["consolidate", "consol", "total"])
FINAL_KW = frozenset(["final", "trial balance", "tb"])
STAGING_KW = frozenset([
    "aje", "adjusting", "adjustments", "elimination",
    "mapping", "bridge", "rollforward", "support", "schedule"
])

CODE_PATTERN = re.compile(r"^[A-Za-z0-9._/-]+$")
SHEET_REF_QUOTED = re.compile(r"'([^']+)'!")
SHEET_REF_BARE = re.compile(r"(?<![A-Za-z0-9_])([A-Za-z0-9 _.-]+)!")

def _safe_str(v):
    return str(v or "").strip()

def _norm(v):
    return _safe_str(v).lower()

def _looks_numeric(v: str) -> bool:
    v = _safe_str(v).replace(",", "")
    if not v:
        return False
    try:
        float(v)
        return True
    except Exception:
        return False

def _extract_sheet_refs_from_formula(formula: str, known_sheets: set[str]) -> list[str]:
    refs = set()
    for s in SHEET_REF_QUOTED.findall(formula):
        if s in known_sheets:
            refs.add(s)
    for s in SHEET_REF_BARE.findall(formula):
        s = s.strip()
        if s in known_sheets:
            refs.add(s)
    return sorted(refs)

def _column_letter(idx: int) -> str:
    out = []
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        out.append(chr(65 + rem))
    return "".join(reversed(out))

def _repetition_ratio(values: list[str]) -> float:
    vals = [_safe_str(v) for v in values if _safe_str(v)]
    if not vals:
        return 1.0
    counts = Counter(vals)
    unique = len(counts)
    total = len(vals)
    return 1.0 - (unique / max(total, 1))

def _best_code_column(columns: dict[int, list[str]]) -> tuple[int | None, dict]:
    best_idx = None
    best_meta = {
        "repetition_ratio": 1.0,
        "valid_count": 0,
    }
    for idx, vals in columns.items():
        clean = [_safe_str(v) for v in vals if _safe_str(v)]
        if len(clean) < 5:
            continue
        valid = [v for v in clean if CODE_PATTERN.match(v) and " " not in v and len(v) <= 32]
        if len(valid) < max(5, int(len(clean) * 0.7)):
            continue
        rr = _repetition_ratio(valid)
        meta = {
            "repetition_ratio": rr,
            "valid_count": len(valid),
        }
        if (
            best_idx is None
            or rr < best_meta["repetition_ratio"]
            or (rr == best_meta["repetition_ratio"] and len(valid) > best_meta["valid_count"])
        ):
            best_idx = idx
            best_meta = meta
    return best_idx, best_meta

def _best_description_column(columns: dict[int, list[str]], code_idx: int | None) -> tuple[int | None, dict]:
    best_idx = None
    best_meta = {
        "repetition_ratio": 1.0,
        "distance_to_code": 10**9,
        "valid_count": 0,
    }
    for idx, vals in columns.items():
        clean = [_safe_str(v) for v in vals if _safe_str(v)]
        if len(clean) < 5:
            continue
        texty = [v for v in clean if not _looks_numeric(v)]
        if len(texty) < max(5, int(len(clean) * 0.6)):
            continue
        rr = _repetition_ratio(texty)
        dist = abs(idx - code_idx) if code_idx is not None else 10**6
        meta = {
            "repetition_ratio": rr,
            "distance_to_code": dist,
            "valid_count": len(texty),
        }
        if (
            best_idx is None
            or rr < best_meta["repetition_ratio"]
            or (rr == best_meta["repetition_ratio"] and dist < best_meta["distance_to_code"])
            or (
                rr == best_meta["repetition_ratio"]
                and dist == best_meta["distance_to_code"]
                and len(texty) > best_meta["valid_count"]
            )
        ):
            best_idx = idx
            best_meta = meta
    return best_idx, best_meta

def _best_final_columns(
    headers: list[str],
    columns: dict[int, list[str]],
    known_upstream_target_cols: set[int] | None = None,
) -> tuple[list[int], dict]:
    candidates = []
    meta = {}
    known_upstream_target_cols = known_upstream_target_cols or set()

    for idx, vals in columns.items():
        clean = [_safe_str(v) for v in vals if _safe_str(v)]
        if len(clean) < 5:
            continue
        numeric_ratio = sum(1 for v in clean if _looks_numeric(v)) / max(len(clean), 1)
        header_l = _norm(headers[idx - 1]) if idx - 1 < len(headers) else ""
        header_hit = int(any(kw in header_l for kw in FINAL_KW))
        upstream_hit = int(idx in known_upstream_target_cols)
        score = (
            numeric_ratio * 1.0
            + header_hit * 1.0
            + upstream_hit * 1.5
        )
        if numeric_ratio >= 0.6 or header_hit or upstream_hit:
            candidates.append((idx, score))
            meta[idx] = {
                "numeric_ratio": numeric_ratio,
                "header_hit": bool(header_hit),
                "upstream_hit": bool(upstream_hit),
            }

    candidates.sort(key=lambda x: x[1], reverse=True)
    return [idx for idx, _ in candidates], meta

def extract_signals(file_path: str, sheet_name: str, wb_full) -> dict:
    """
    Open sheet with formulas, extract Layer-0 features, and compute Layer-1 signals.
    """
    ws_meta = wb_full[sheet_name]
    hidden = ws_meta.sheet_state in ("hidden", "veryHidden")
    known_sheets = set(wb_full.sheetnames)

    wb_f = load_workbook(file_path, read_only=True, data_only=False)
    ws = wb_f[sheet_name]

    headers = []
    formulas = []
    refs = []
    flat = []
    rows_cache = []
    columns = {}

    max_rows = 250
    max_flat = 300

    for r_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
        row_vals = []
        for c_idx, cell in enumerate(row, 1):
            raw = cell.value
            v = _safe_str(raw)
            row_vals.append(v)
            columns.setdefault(c_idx, []).append(v)
            if v:
                flat.append(v.lower())
                if isinstance(raw, str) and v.startswith("="):
                    formulas.append(v)
                    refs.extend(_extract_sheet_refs_from_formula(v, known_sheets))
        if r_idx == 1:
            headers = [v for v in row_vals if v]
        if 2 <= r_idx <= 6:
            rows_cache.append(row_vals)
        if r_idx >= max_rows or len(flat) >= max_flat:
            break

    wb_f.close()

    flat_set = set(flat)
    hdr_lower = [_norm(h) for h in headers]
    coa_hits = sum(1 for s in REQUIRED_SECTIONS if s in flat_set)

    # candidate columns
    code_idx, code_meta = _best_code_column(columns)
    desc_idx, desc_meta = _best_description_column(columns, code_idx)
    final_candidates, final_meta = _best_final_columns(headers, columns, set())

    company_header_hit = any(any(kw in h for kw in COMPANY_KW) for h in hdr_lower)
    aje_header_hit = any(any(kw in h for kw in AJE_KW) for h in hdr_lower)
    consol_header_hit = any(any(kw in h for kw in CONSOL_KW) for h in hdr_lower)
    staging_header_hit = any(any(kw in h for kw in STAGING_KW) for h in hdr_lower)

    return {
        # raw extraction
        "_sheet_name": sheet_name,
        "_headers": headers,
        "_sample_rows": rows_cache,
        "_all_values_flat": flat[:300],
        "_formulas": formulas[:30],
        "_outgoing_refs": sorted(set(refs)),
        "_candidate_code_columns": [code_idx] if code_idx else [],
        "_candidate_desc_columns": [desc_idx] if desc_idx else [],
        "_candidate_final_columns": final_candidates[:5],
        "_code_meta": code_meta,
        "_desc_meta": desc_meta,
        "_final_meta": final_meta,

        # Layer 1
        "COA_SIGNAL": int(coa_hits >= 3),
        "FORMULA_SIGNAL": int(bool(formulas)),
        "CROSS_REF_SIGNAL": int(bool(refs)),
        "REFERENCED_BY_SIGNAL": 0,  # filled later after graph build
        "COMPANY_COLUMN_SIGNAL": int(company_header_hit),
        "AJE_SIGNAL": int(aje_header_hit),
        "CONSOLIDATE_SIGNAL": int(consol_header_hit),
        "HAS_CODE_COLUMN": int(code_idx is not None),
        "HAS_DESCRIPTION_COLUMN": int(desc_idx is not None),
        "HAS_FINAL_COLUMN": int(bool(final_candidates)),
        "FINAL_REFERENCE_SIGNAL": 0,  # filled later after graph+target analysis
        "TB_REFERENCE_SIGNAL": 0,     # filled later after graph build
        "STAGING_ROLE_SIGNAL": int(staging_header_hit),
        "HIDDEN_SIGNAL": int(hidden),

        # backward-compatible aliases
        "CODE_COLUMN_SIGNAL": int(code_idx is not None),
        "FINAL_COLUMN_SIGNAL": int(bool(final_candidates)),

        # metadata for downstream layers
        "_coa_hits": coa_hits,
    }


# ── Layer 2: pattern computation ─────────────────────────────────────────────
def compute_patterns(sig: dict) -> dict:
    fs = (
        sig["COA_SIGNAL"]
        and sig["FORMULA_SIGNAL"]
        and sig["CROSS_REF_SIGNAL"]
        and sig["COMPANY_COLUMN_SIGNAL"]
        and not sig["HIDDEN_SIGNAL"]
        and not sig["HAS_CODE_COLUMN"]
    )

    partial = (
        sig["COA_SIGNAL"]
        and (sig["FORMULA_SIGNAL"] or sig["CROSS_REF_SIGNAL"])
        and not sig["HIDDEN_SIGNAL"]
    )

    tb = (
        sig["HAS_CODE_COLUMN"]
        and sig["HAS_DESCRIPTION_COLUMN"]
        and sig["HAS_FINAL_COLUMN"]
        and not sig["HIDDEN_SIGNAL"]
    )

    strong_tb = (
        tb
        and (
            sig["FINAL_REFERENCE_SIGNAL"]
            or sig["TB_REFERENCE_SIGNAL"]
            or sig["REFERENCED_BY_SIGNAL"]
        )
    )

    staging = (
        sig["STAGING_ROLE_SIGNAL"]
        and (sig["AJE_SIGNAL"] or sig["FORMULA_SIGNAL"] or sig["CROSS_REF_SIGNAL"])
    )

    return {
        "FS_PATTERN": int(bool(fs)),
        "TB_PATTERN": int(bool(tb)),
        "PARTIAL_FS_PATTERN": int(bool(partial)),
        "STRONG_TB_PATTERN": int(bool(strong_tb)),
        "STAGING_PATTERN": int(bool(staging)),
    }


# ── Layer 3: dependency graph + source/staging logic ────────────────────────
def build_graph(all_signals: dict) -> dict:
    graph = {
        sn: {
            "outgoing_refs": list(all_signals[sn].get("_outgoing_refs", [])),
            "incoming_refs": [],
            "role_in_graph": "UNKNOWN",
            "consolidate": False,
            "attention_boost": False,
            "aje_source_role": False,
            "path_to_tb": [],
            "path_valid": False,
        }
        for sn in all_signals
    }

    # populate incoming refs
    for sn, data in graph.items():
        for ref in data["outgoing_refs"]:
            if ref in graph:
                graph[ref]["incoming_refs"].append(sn)

    # fill REFERENCED_BY and TB_REFERENCE signals
    for sn, data in graph.items():
        if data["incoming_refs"]:
            all_signals[sn]["REFERENCED_BY_SIGNAL"] = 1
            all_signals[sn]["TB_REFERENCE_SIGNAL"] = 1

    # first pass roles
    for sn, data in graph.items():
        sig = all_signals[sn]
        pat = compute_patterns(sig)

        has_out = bool(data["outgoing_refs"])
        has_in = bool(data["incoming_refs"])

        if pat["TB_PATTERN"] or pat["STRONG_TB_PATTERN"]:
            data["role_in_graph"] = "TB"
        elif pat["STAGING_PATTERN"]:
            data["role_in_graph"] = "STAGING"
        elif pat["FS_PATTERN"] or (has_out and not has_in):
            data["role_in_graph"] = "FS"
        elif has_out and has_in:
            data["role_in_graph"] = "INTERMEDIATE"
        elif has_in and not has_out:
            data["role_in_graph"] = "TB"
        else:
            data["role_in_graph"] = "UNKNOWN"

    # detect staging source role (AJE-like supporter)
    for sn, data in graph.items():
        sig = all_signals[sn]
        if sig["AJE_SIGNAL"] or sig["STAGING_ROLE_SIGNAL"]:
            # if sheet is primarily upstream/supportive, treat as staging source
            if data["incoming_refs"] and not sig["COA_SIGNAL"]:
                data["aje_source_role"] = True

    # active sheet boost
    active_title = wb_full.active.title if "wb_full" in globals() and wb_full.active else None
    for sn, data in graph.items():
        sig = all_signals[sn]
        pat = compute_patterns(sig)
        if active_title == sn and pat["FS_PATTERN"]:
            data["attention_boost"] = True

    # simple path-to-tb discovery: FS -> TB or FS -> INTERMEDIATE/STAGING -> TB
    for sn, data in graph.items():
        sig = all_signals[sn]
        pat = compute_patterns(sig)

        if not (pat["FS_PATTERN"] or pat["PARTIAL_FS_PATTERN"]):
            continue

        # direct
        for ref in data["outgoing_refs"]:
            if ref in graph and graph[ref]["role_in_graph"] == "TB":
                data["path_to_tb"] = [sn, ref]
                data["path_valid"] = True
                break

        if data["path_valid"]:
            continue

        # indirect
        for mid in data["outgoing_refs"]:
            if mid not in graph:
                continue
            if graph[mid]["role_in_graph"] not in ("INTERMEDIATE", "STAGING"):
                continue
            for ref2 in graph[mid]["outgoing_refs"]:
                if ref2 in graph and graph[ref2]["role_in_graph"] == "TB":
                    data["path_to_tb"] = [sn, mid, ref2]
                    data["path_valid"] = True
                    break
            if data["path_valid"]:
                break

    return graph


# ── Layer 4: hard gates for main-sheet candidacy ────────────────────────────
def apply_gates(sig: dict, pat: dict, graph_node: dict) -> dict:
    if sig["HIDDEN_SIGNAL"]:
        return {"passed": False, "blocked_by": "GATE_1"}

    consolidate_exempt = graph_node.get("consolidate", False)
    if (
        not sig["COMPANY_COLUMN_SIGNAL"]
        and not sig["CONSOLIDATE_SIGNAL"]
        and not consolidate_exempt
    ):
        return {"passed": False, "blocked_by": "GATE_2"}

    if pat["TB_PATTERN"] or pat["STRONG_TB_PATTERN"]:
        return {"passed": False, "blocked_by": "GATE_3"}

    if (
        graph_node["role_in_graph"] == "TB"
        and not pat["FS_PATTERN"]
        and not pat["PARTIAL_FS_PATTERN"]
    ):
        return {"passed": False, "blocked_by": "GATE_4"}

    if pat["STAGING_PATTERN"] and (
        graph_node["role_in_graph"] == "STAGING"
        or graph_node.get("aje_source_role", False)
    ):
        return {"passed": False, "blocked_by": "GATE_5"}

    return {"passed": True, "blocked_by": None}


# ── Layer 5: technical main-sheet confidence (true softmax) ─────────────────
def signal_strength(sig: dict, pat: dict, attention_boost: bool = False) -> float:
    return (
        pat["FS_PATTERN"] * 2.40
        + pat["PARTIAL_FS_PATTERN"] * 1.10
        + sig["CROSS_REF_SIGNAL"] * 0.45
        + sig["COMPANY_COLUMN_SIGNAL"] * 0.40
        + sig["CONSOLIDATE_SIGNAL"] * 0.35
        + (0.20 if attention_boost else 0.0)
    )

def softmax_confidence(strengths: dict, tau: float = 0.75) -> dict:
    if not strengths:
        return {}
    logits = {k: math.exp(v / tau) for k, v in strengths.items()}
    total = sum(logits.values()) or 1e-9
    return {s: round(v / total, 4) for s, v in logits.items()}


# ── Layer 6: business/presentation main-sheet arbitration ───────────────────
def classify_sheet_type(sheet_name: str, sig: dict, pat: dict, graph_node: dict, title: str = "") -> str:
    name_l = _norm(sheet_name)
    title_l = _norm(title)
    combined = f"{name_l} || {title_l}"

    strong_fs_kws = [
        "balance sheet", "balance sheets",
        "statement of operations", "statement of income",
        "profit and loss", "p&l",
        "cash flow", "cash flows", "statement of cash flows",
        "change in equity", "stockholders' equity",
        "financial statements", "report"
    ]

    if pat["TB_PATTERN"] or pat["STRONG_TB_PATTERN"] or graph_node["role_in_graph"] == "TB":
        return "SOURCE_TB"

    if (
        pat["STAGING_PATTERN"]
        or graph_node.get("aje_source_role", False)
        or graph_node["role_in_graph"] == "STAGING"
    ):
        return "ADJUSTMENT_STAGING"

    if pat["FS_PATTERN"]:
        return "REPORTING_FS"

    if any(k in combined for k in strong_fs_kws) and sig["COA_SIGNAL"]:
        return "REPORTING_FS"

    if graph_node.get("consolidate", False) or graph_node["role_in_graph"] == "INTERMEDIATE":
        return "INTERMEDIATE_CONSOLIDATION"

    if sig["COA_SIGNAL"] and not sig["CROSS_REF_SIGNAL"]:
        return "AUXILIARY_SCHEDULE"

    return "UNKNOWN"

def presentation_rank(sheet_name: str, sig: dict, pat: dict, graph_node: dict, gate: dict, conf: float, title: str = "") -> tuple:
    stype = classify_sheet_type(sheet_name, sig, pat, graph_node, title)
    blocked_by = gate.get("blocked_by")
    dq = "NONE" if gate.get("passed") and not blocked_by else ("TECHNICAL" if blocked_by == "GATE_2" and stype == "REPORTING_FS" else "CRITICAL")

    name_l = _norm(sheet_name)
    title_l = _norm(title)
    combined = f"{name_l} || {title_l}"

    canonical_title = int(any(k in combined for k in [
        "balance sheet", "balance sheets",
        "statement of operations", "statement of income",
        "profit and loss", "p&l",
        "cash flow", "cash flows", "statement of cash flows",
        "change in equity", "stockholders' equity",
        "financial statements", "report"
    ]))
    final_output = int(
        stype == "REPORTING_FS"
        and sig["COA_SIGNAL"]
        and not pat["TB_PATTERN"]
        and not pat["STAGING_PATTERN"]
    )
    presentation_layout = int(
        sig["COA_SIGNAL"]
        and not sig["HAS_CODE_COLUMN"]
        and graph_node["role_in_graph"] != "TB"
    )
    non_critical = int(dq in ("NONE", "TECHNICAL"))
    return (
        non_critical,
        final_output,
        canonical_title,
        presentation_layout,
        int(stype == "REPORTING_FS"),
        conf,
    )


# ── Layer 7: TB/card-sheet validation ───────────────────────────────────────
def tb_rank(sheet_name: str, sig: dict, pat: dict, graph_node: dict) -> tuple:
    return (
        int(sig["HIDDEN_SIGNAL"] == 0),
        pat["STRONG_TB_PATTERN"],
        pat["TB_PATTERN"],
        sig["HAS_CODE_COLUMN"],
        sig["HAS_DESCRIPTION_COLUMN"],
        sig["HAS_FINAL_COLUMN"],
        sig["FINAL_REFERENCE_SIGNAL"],
        sig["TB_REFERENCE_SIGNAL"],
        sig["REFERENCED_BY_SIGNAL"],
        int(graph_node.get("path_valid", False)),
    )
'''


_CODER_OUTPUT_SCHEMA = """
REQUIRED OUTPUT — Return a single JSON object, no markdown:
{
  "main_sheet_exists": true/false,
  "main_sheet_name": "<sheet name or null>",
  "is_card_sheet": "<TB sheet name or null>",
  "technical_main_sheet": "<sheet name or null>",
  "presentation_main_sheet": "<sheet name or null>",
  "technical_tb_sheet": "<sheet name or null>",
  "main_sheet_confirmed": true/false,
  "confidence": 0.0,
  "decision_mode": "technical_default|business_override|business_override_with_tb_validation|no_valid_sheet",
  "nn_evidence": {
    "layer1": {
      "COA_SIGNAL": 0,
      "FORMULA_SIGNAL": 0,
      "CROSS_REF_SIGNAL": 0,
      "REFERENCED_BY_SIGNAL": 0,
      "COMPANY_COLUMN_SIGNAL": 0,
      "AJE_SIGNAL": 0,
      "CONSOLIDATE_SIGNAL": 0,
      "HAS_CODE_COLUMN": 0,
      "HAS_DESCRIPTION_COLUMN": 0,
      "HAS_FINAL_COLUMN": 0,
      "FINAL_REFERENCE_SIGNAL": 0,
      "TB_REFERENCE_SIGNAL": 0,
      "STAGING_ROLE_SIGNAL": 0,
      "HIDDEN_SIGNAL": 0
    },
    "layer2": {
      "FS_PATTERN": 0,
      "TB_PATTERN": 0,
      "PARTIAL_FS_PATTERN": 0,
      "STRONG_TB_PATTERN": 0,
      "STAGING_PATTERN": 0
    },
    "layer3": {
      "outgoing_refs": [],
      "incoming_refs": [],
      "role_in_graph": "FS|TB|INTERMEDIATE|STAGING|UNKNOWN",
      "consolidate": false,
      "attention_boost": false,
      "aje_source_role": false,
      "path_to_tb": [],
      "path_valid": false
    },
    "layer4": {
      "passed": true,
      "blocked_by": "GATE_1|GATE_2|GATE_3|GATE_4|GATE_5|null"
    },
    "layer5_confidence": 0.0,
    "all_sheets_confidence": {"<sheet>": 0.0}
  },
  "hidden_sheets": [],
  "tb_sheets": [],
  "relationship": {
    "main_to_tb_path": [],
    "path_valid": false
  },
  "reasoning": "<one concise English sentence>"
}
"""


def build_coder_system_prompt() -> str:
    """
    ORC Coder Agent system prompt.

    ReAct behavior + 7-layer Neural PROMAT.
    The coder may generate Python to inspect workbook structure and return
    NN-consistent JSON.
    """
    return f"""
{FULL_NN_PROMAT}

╔══════════════════════════════════════════════════════════╗
║  AGENT ROLE: Coder Agent (ReAct + Neural PROMAT)        ║
╚══════════════════════════════════════════════════════════╝

You are the Coder Agent in a LangGraph ORC pipeline.
Use ReAct behavior:
  Reason → generate code → execute → observe → repeat.

CORE CONSTRAINTS
────────────────
• Generate code only when it adds new information not already observed.
• Every code path must follow the 7-layer Neural PROMAT.
• Do not trust sheet names alone.
• Do not trust highlighted titles alone.
• The final output must include:
    • main sheet
    • TB/card sheet
    • validated relationship
• Use true softmax logic, not simple linear ratio normalisation.
• Return JSON only.

PYTHON HELPERS (use these directly in generated code)
──────────────────────────────────────────────────────
{_NN_PYTHON_HELPERS}

EXECUTION ORDER FOR GENERATED CODE
──────────────────────────────────
1. Load workbook (read_only mode for inspection, formulas visible).
2. Enumerate sheets and visibility.
   For hidden sheets:
     • HIDDEN_SIGNAL = 1
     • GATE_1 for main-sheet candidacy
3. For each visible sheet:
     • run extract_signals(...)
4. For each visible sheet:
     • run compute_patterns(...)
5. After all sheets are scanned:
     • run build_graph(all_signals)
6. For each sheet:
     • run apply_gates(sig, pat, graph_node)
7. For passing main-sheet candidates:
     • compute signal_strength(...)
     • compute true softmax with softmax_confidence(...)
     • identify technical main-sheet winner
8. Run Layer 6:
     • classify technical main sheet
     • identify presentation/main override candidate
     • determine final main_sheet_name
9. Run Layer 7:
     • rank TB candidates with tb_rank(...)
     • choose technical_tb_sheet / is_card_sheet
     • preserve relationship.main_to_tb_path and relationship.path_valid
10. If decision_mode = business_override and TB relationship is valid:
      upgrade to business_override_with_tb_validation
11. Return one JSON object matching the required schema.

IMPORTANT IMPLEMENTATION RULES
──────────────────────────────
• A TB/card sheet must never become the final main reporting sheet.
• A staging/AJE-support sheet must never become the final main reporting sheet.
• GATE_2 is technical only.
• GATE_1 / GATE_3 / GATE_4 / GATE_5 are critical.
• HAS_FINAL_COLUMN may be determined by:
    • header text
    • upstream formula targeting
    • numeric ending-balance behavior
• If only one strong FINAL column exists, treat it as FINAL.
• If there is no AJE, treat AJE = 0.
• If only FINAL exists, FINAL remains the effective amount source.

{_CODER_OUTPUT_SCHEMA}
""".strip()