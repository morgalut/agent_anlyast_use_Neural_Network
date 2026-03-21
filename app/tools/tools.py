"""
tools.py — High-performance Excel toolkit for ORC pipeline agents.

Design principles
─────────────────
• read_only=True  everywhere except formula extraction (saves 10–40× memory).
• data_only=False only when callers explicitly need formulas.
• Chunk-based scanning  — never pull an entire multi-MB sheet into memory at once.
• Single-open helpers   — open / close the workbook once per top-level call.
• Lazy column-letter cache — avoid repeated openpyxl util calls.
• All public functions return plain dicts / lists (JSON-serialisable).
"""

from __future__ import annotations

import re
from datetime import datetime
from functools import lru_cache
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# ─────────────────────────────────────────────────────────────────────────────
#  Constants
# ─────────────────────────────────────────────────────────────────────────────

_CHUNK_ROWS        = 500    # rows per scanning chunk (memory ceiling)
_PREVIEW_ROWS      = 6      # rows returned in preview
_PREVIEW_COLS      = 8      # cols returned in preview
_HEADER_SCAN_ROWS  = 20     # how many rows to inspect for the header
_HEADER_SCAN_COLS  = 16
_TITLE_SCAN_ROWS   = 8
_TITLE_SCAN_COLS   = 8
_FORMULA_CAP       = 30     # max formula samples per sheet
_FLAT_CAP          = 300    # max flat values per sheet

# COA section keywords (used by every agent)
_COA_SECTIONS = frozenset(
    [
        "assets", "current assets", "long-term assets",
        "liabilities and equity", "liabilities & equity",
        "current liabilities", "long-term liabilities", "equity",
    ]
)

_COMPANY_CURRENCY_KEYWORDS = frozenset(
    ["nis", "dollar", "$", "usd", "ils", "inc", "ltd", "eur", "gbp"]
)

_CONSOLIDATE_KEYWORDS = frozenset(["consolidate", "consol", "total"])

_AJE_KEYWORDS = frozenset(["aje", "adjusting", "journal entry", "journal entries"])

_TB_HEADERS = frozenset(["final", "tb", "trial balance", "debit", "credit"])

_OUTPUT_KEYWORDS = [
    "fs", "financial statements", "report", "reports", "summary",
    "actual vs budget", "cash flow", "cf", "shareholders equity",
    "equity", "p&l", "income statement", "balance sheet",
]

_SOURCE_KEYWORDS = [
    "accounts", "gl", "aje", "tb", "wp", "intercompany",
    "captable", "severence", "translation", "שער", "ש\"ח",
]


# ─────────────────────────────────────────────────────────────────────────────
#  Low-level helpers
# ─────────────────────────────────────────────────────────────────────────────

@lru_cache(maxsize=512)
def _col_letter(n: int) -> str:
    return get_column_letter(n)


def _norm(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _is_meaningful(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _looks_numeric(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return True
    if value is None:
        return False
    try:
        float(str(value).strip().replace(",", ""))
        return True
    except ValueError:
        return False


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _sheet_refs_from_formula(formula: str) -> list[str]:
    """Extract sheet names referenced inside a formula, e.g. ='TB USD'!B5 → ['TB USD']."""
    return re.findall(r"'([^']+)'!", formula) + re.findall(r"!([A-Za-z0-9_]+)!", formula)


# ─────────────────────────────────────────────────────────────────────────────
#  Workbook open helpers  (read_only whenever possible)
# ─────────────────────────────────────────────────────────────────────────────

def _open_ro(file_path: str) -> Workbook:
    """Open in read-only mode — fastest, minimal memory."""
    return load_workbook(filename=file_path, read_only=True, data_only=True)


def _open_formulas(file_path: str) -> Workbook:
    """Open with formula strings (data_only=False), still read-only."""
    return load_workbook(filename=file_path, read_only=True, data_only=False)


def _open_full(file_path: str) -> Workbook:
    """Open fully — only for operations that need cell styles / visibility."""
    return load_workbook(filename=file_path, data_only=True)


def _visibility(wb: Workbook, sheet_name: str) -> str:
    """Return 'visible', 'hidden', or 'veryHidden'."""
    try:
        state = wb.get_sheet_visibility(sheet_name)
        return state or "visible"
    except Exception:
        return "visible"


# ─────────────────────────────────────────────────────────────────────────────
#  Fast sheet scanning  (chunked, streaming)
# ─────────────────────────────────────────────────────────────────────────────

def _scan_used_area_fast(ws) -> dict:
    """
    One-pass scan of all cells in read_only mode.
    Returns the bounding box + non-empty count.
    Never loads all rows at once — processes _CHUNK_ROWS at a time.
    """
    min_row = min_col = None
    max_row = max_col = 0
    non_empty = 0

    for row in ws.iter_rows():
        for cell in row:
            if _is_meaningful(cell.value):
                non_empty += 1
                r, c = cell.row, cell.column
                if min_row is None or r < min_row:
                    min_row = r
                if min_col is None or c < min_col:
                    min_col = c
                if r > max_row:
                    max_row = r
                if c > max_col:
                    max_col = c

    if min_row is None:
        return {
            "has_content": False,
            "min_row": None, "max_row": None,
            "min_col": None, "max_col": None,
            "min_col_letter": None, "max_col_letter": None,
            "non_empty_count": 0, "used_rows": 0, "used_cols": 0,
        }

    return {
        "has_content": True,
        "min_row": min_row, "max_row": max_row,
        "min_col": min_col, "max_col": max_col,
        "min_col_letter": _col_letter(min_col),
        "max_col_letter": _col_letter(max_col),
        "non_empty_count": non_empty,
        "used_rows": max_row - min_row + 1,
        "used_cols": max_col - min_col + 1,
    }


def _extract_title_fast(ws, max_r: int = _TITLE_SCAN_ROWS, max_c: int = _TITLE_SCAN_COLS) -> dict:
    """Scan the top-left region for the best title candidate."""
    best = ""
    best_row = best_col = None
    cap_r = min(getattr(ws, "max_row", max_r) or max_r, max_r)
    cap_c = min(getattr(ws, "max_column", max_c) or max_c, max_c)

    for row in ws.iter_rows(min_row=1, max_row=cap_r, min_col=1, max_col=cap_c, values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            txt = str(v).strip()
            if len(txt) < 3 or _looks_numeric(v):
                continue
            if len(txt) > len(best):
                best = txt
                best_row = cell.row
                best_col = cell.column

    return {"title": best, "row": best_row, "col": best_col}


def _detect_header_row_fast(ws) -> dict:
    """
    Score each of the first _HEADER_SCAN_ROWS rows as a potential header.
    Higher string density + more unique values = better header.
    """
    best_row = best_score = None
    best_values: list = []

    cap_r = min(getattr(ws, "max_row", _HEADER_SCAN_ROWS) or _HEADER_SCAN_ROWS, _HEADER_SCAN_ROWS)
    cap_c = min(getattr(ws, "max_column", _HEADER_SCAN_COLS) or _HEADER_SCAN_COLS, _HEADER_SCAN_COLS)

    for row in ws.iter_rows(min_row=1, max_row=cap_r, min_col=1, max_col=cap_c, values_only=True):
        non_empty   = [v for v in row if _is_meaningful(v)]
        string_cnt  = sum(1 for v in non_empty if isinstance(v, str))
        numeric_cnt = sum(1 for v in non_empty if _looks_numeric(v))
        unique_str  = len({str(v).strip().lower() for v in non_empty if isinstance(v, str)})

        score = string_cnt * 3 + unique_str - numeric_cnt * 2

        if score > (best_score or -1) and string_cnt >= 2:
            best_score = score
            best_row   = row[0]  # placeholder — we need real row index
            best_values = list(row)

    # iter_rows values_only doesn't give row number — re-iterate with index
    actual_best_row = None
    if best_values:
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=cap_r, min_col=1, max_col=cap_c, values_only=True),
            start=1,
        ):
            if list(row) == best_values:
                actual_best_row = r_idx
                break

    return {
        "header_row":    actual_best_row,
        "header_values": best_values,
        "header_score":  best_score or 0,
    }


def _preview_fast(ws, used: dict, rows: int = _PREVIEW_ROWS, cols: int = _PREVIEW_COLS) -> list[list]:
    """Stream preview rows from the detected used area — no full-sheet load."""
    if not used["has_content"]:
        return []
    min_r = used["min_row"]
    min_c = used["min_col"]
    max_r = min(used["max_row"], min_r + rows - 1)
    max_c = min(used["max_col"], min_c + cols - 1)

    return [
        list(row)
        for row in ws.iter_rows(
            min_row=min_r, max_row=max_r,
            min_col=min_c, max_col=max_c,
            values_only=True,
        )
    ]


def _extract_ocr_snapshot(ws_data, ws_formula) -> dict:
    """
    OCR-style extraction using two worksheet handles:
      ws_data    — values (data_only=True)   for headers + sample rows
      ws_formula — formulas (data_only=False) for formula references

    Returns:
      headers, sample_rows, formula_samples, formula_sheet_refs,
      coa_sections_found, company_columns, aje_found, consolidate_found,
      has_tb_signature, all_values_flat
    """
    headers: list[str] = []
    sample_rows: list[list] = []
    formula_samples: list[str] = []
    formula_sheet_refs: list[str] = []
    all_values_flat: list[str] = []

    for row_idx, row in enumerate(ws_formula.iter_rows(values_only=False), start=1):
        row_vals: list = []
        for cell in row:
            val = cell.value
            str_val = str(val).strip() if val is not None else ""

            if str_val:
                all_values_flat.append(str_val)
                if _is_formula(str_val):
                    if len(formula_samples) < _FORMULA_CAP:
                        formula_samples.append(str_val[:120])
                    for ref in _sheet_refs_from_formula(str_val):
                        if ref not in formula_sheet_refs:
                            formula_sheet_refs.append(ref)
            row_vals.append(val)

        if row_idx == 1:
            headers = [str(v).strip() for v in row_vals if _is_meaningful(v)]
        elif row_idx <= 7:
            sample_rows.append(row_vals)

        if len(all_values_flat) >= _FLAT_CAP:
            break

    # ── COA analysis ─────────────────────────────────────────────────────────
    flat_lower = {v.lower() for v in all_values_flat}
    coa_sections_found = [s for s in _COA_SECTIONS if s in flat_lower]

    # ── Company columns ───────────────────────────────────────────────────────
    company_columns: list[str] = []
    for h in headers:
        if any(kw in h.lower() for kw in _COMPANY_CURRENCY_KEYWORDS):
            company_columns.append(h)

    # ── AJE detection ─────────────────────────────────────────────────────────
    aje_found = any(
        any(kw in h.lower() for kw in _AJE_KEYWORDS) for h in headers
    ) or any(any(kw in v.lower() for kw in _AJE_KEYWORDS) for v in all_values_flat[:80])

    # ── CONSOLIDATE detection ─────────────────────────────────────────────────
    consolidate_found = any(
        any(kw in h.lower() for kw in _CONSOLIDATE_KEYWORDS) for h in headers
    )

    # ── TB signature (code + description + FINAL) ─────────────────────────────
    tb_header_hits = sum(1 for h in headers if any(kw in h.lower() for kw in _TB_HEADERS))
    code_pattern   = re.compile(r"^[A-Za-z0-9]+$")
    code_col_vals  = [v for v in all_values_flat[:80] if code_pattern.match(v)]
    has_tb_sig     = tb_header_hits >= 1 and len(code_col_vals) >= 3 and len(coa_sections_found) == 0

    return {
        "headers":             headers,
        "sample_rows":         sample_rows,
        "formula_samples":     formula_samples,
        "formula_sheet_refs":  formula_sheet_refs,
        "all_values_flat":     all_values_flat[:_FLAT_CAP],
        "coa_sections_found":  coa_sections_found,
        "company_columns":     company_columns,
        "aje_found":           aje_found,
        "consolidate_found":   consolidate_found,
        "has_tb_signature":    has_tb_sig,
        "ocr_used":            True,
    }


# ─────────────────────────────────────────────────────────────────────────────
#  Scoring helpers  (shared between inspect_workbook and detect_main_sheet)
# ─────────────────────────────────────────────────────────────────────────────

def _score_sheet(
    sheet_name: str,
    title_norm: str,
    used: dict,
    header_info: dict,
    ocr: dict,
) -> tuple[int, int, list[str]]:
    """
    Return (output_score, source_score, breakdown).
    Incorporates PROMAT-aligned COA / company / AJE / CONSOLIDATE signals.
    """
    sn = sheet_name.strip().lower()
    output_score = 0
    source_score = 0
    breakdown: list[str] = []

    # ── Output / reporting signals ────────────────────────────────────────────
    if sn == "fs":
        output_score += 80; breakdown.append("exact name 'FS'")
    if "financial statements" in title_norm:
        output_score += 70; breakdown.append("title: Financial Statements")
    if "report" in sn or "report" in title_norm:
        output_score += 20; breakdown.append("report keyword")
    if "summary" in sn or "summary" in title_norm:
        output_score += 18; breakdown.append("summary keyword")
    if "actual vs budget" in sn:
        output_score += 18; breakdown.append("actual vs budget")
    if "cash flow" in title_norm or sn == "cf":
        output_score += 16; breakdown.append("cash flow")
    if "shareholders equity" in sn:
        output_score += 16; breakdown.append("equity statement")
    if "slides" in sn:
        output_score += 10; breakdown.append("slides keyword")

    # ── COA PROMAT signals ────────────────────────────────────────────────────
    n_sections = len(ocr["coa_sections_found"])
    output_score += int(n_sections * (20 / 7))
    if n_sections > 0:
        breakdown.append(f"COA sections found: {n_sections}/7")

    # ── Company columns ───────────────────────────────────────────────────────
    if ocr["company_columns"]:
        output_score += 20; breakdown.append(f"company cols: {ocr['company_columns'][:3]}")

    # ── AJE ──────────────────────────────────────────────────────────────────
    if ocr["aje_found"]:
        output_score += 15; breakdown.append("AJE column found")

    # ── CONSOLIDATE ───────────────────────────────────────────────────────────
    if ocr["consolidate_found"]:
        output_score += 15; breakdown.append("CONSOLIDATE column found")

    # ── Formula refs to other sheets (reporting sheet points to TB) ───────────
    if ocr["formula_sheet_refs"]:
        output_score += 10; breakdown.append(f"formula refs: {ocr['formula_sheet_refs'][:3]}")

    # ── Source / support signals ──────────────────────────────────────────────
    if "accounts" in sn:
        source_score += 35; breakdown.append("accounts keyword")
    if sn.startswith("gl"):
        source_score += 30; breakdown.append("general ledger")
    if sn.startswith("aje"):
        source_score += 26; breakdown.append("AJE sheet name")
    if sn.startswith("tb"):
        source_score += 20; breakdown.append("trial balance")
    if "intercompany" in sn:
        source_score += 20; breakdown.append("intercompany")
    if "captable" in sn:
        source_score += 16; breakdown.append("cap table")
    if "wp" in sn:
        source_score += 14; breakdown.append("working paper")
    if "severence" in sn:
        source_score += 10; breakdown.append("support schedule")
    if "ש\"ח" in sn or "שער" in title_norm:
        source_score += 10; breakdown.append("FX/translation")

    # ── TB signature penalises output score ───────────────────────────────────
    if ocr["has_tb_signature"]:
        output_score -= 30; breakdown.append("TB-like signature (penalty)")
        source_score += 20; breakdown.append("TB-like signature (bonus)")

    # ── Structural hints ──────────────────────────────────────────────────────
    if used["has_content"]:
        if used.get("min_row", 99) <= 5 and title_norm:
            output_score += 8; breakdown.append("early title")
        if header_info.get("header_row") is not None and header_info["header_row"] <= 5:
            source_score += 6; breakdown.append("early header row")
        if used["non_empty_count"] >= 30:
            source_score += 4; breakdown.append("dense sheet")

    final_output = output_score - int(source_score * 0.15)
    return final_output, source_score, breakdown


# ─────────────────────────────────────────────────────────────────────────────
#  ExcelNavigator  (chunked, read-only navigation for agent tool calls)
# ─────────────────────────────────────────────────────────────────────────────

class ExcelNavigator:
    """
    Stateful Excel navigator backed by an openpyxl read_only workbook.
    All large reads are chunked to _CHUNK_ROWS rows at a time.
    """

    def __init__(self, file_path: str, sheet_name: str | None = None, data_only: bool = True):
        self.file_path = file_path
        self._data_only = data_only
        self.workbook: Workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=data_only,
        )
        self.sheet = (
            self.workbook[sheet_name]
            if sheet_name
            else self.workbook.active
        )
        self.current_row = 1
        self.current_col = 1

    # ── Metadata ──────────────────────────────────────────────────────────────

    def workbook_info(self) -> dict:
        return {
            "file_path":   self.file_path,
            "sheet_names": self.workbook.sheetnames,
            "active_sheet": self.sheet.title,
            "max_row":     self.sheet.max_row,
            "max_column":  self.sheet.max_column,
            "dimensions":  self.sheet.dimensions,
        }

    def list_sheets(self) -> list[str]:
        return list(self.workbook.sheetnames)

    def use_sheet(self, sheet_name: str) -> str:
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. "
                f"Available: {self.workbook.sheetnames}"
            )
        self.sheet = self.workbook[sheet_name]
        self.current_row = self.current_col = 1
        return self.sheet.title

    # ── Pointer ───────────────────────────────────────────────────────────────

    def set_position(self, row: int, col: int) -> dict:
        self._validate(row, col)
        self.current_row = row
        self.current_col = col
        return self.position()

    def position(self) -> dict:
        return {
            "row": self.current_row,
            "col": self.current_col,
            "col_letter": _col_letter(self.current_col),
            "cell": f"{_col_letter(self.current_col)}{self.current_row}",
        }

    def move(self, row_offset: int = 0, col_offset: int = 0) -> dict:
        nr = self.current_row + row_offset
        nc = self.current_col + col_offset
        self._validate(nr, nc)
        self.current_row = nr
        self.current_col = nc
        return self.position()

    # ── Cell / row / column reads ─────────────────────────────────────────────

    def cell(self, row: int | None = None, col: int | str | None = None) -> dict:
        r = row if row is not None else self.current_row
        c = self._nc(col) if col is not None else self.current_col
        self._validate(r, c)
        # In read_only mode, fastest way to read a single cell is iter_rows
        for row_cells in self.sheet.iter_rows(
            min_row=r, max_row=r, min_col=c, max_col=c, values_only=False
        ):
            cell_obj = row_cells[0]
            return {
                "sheet":      self.sheet.title,
                "row":        r,
                "col":        c,
                "col_letter": _col_letter(c),
                "coordinate": cell_obj.coordinate,
                "value":      cell_obj.value,
            }
        return {"sheet": self.sheet.title, "row": r, "col": c, "value": None}

    def current_cell(self) -> dict:
        return self.cell()

    def row_values(
        self,
        row: int | None = None,
        start_col: int = 1,
        end_col: int | None = None,
    ) -> dict:
        r = row if row is not None else self.current_row
        ec = end_col or self.sheet.max_column or start_col
        self._validate(r, start_col)
        values = next(
            self.sheet.iter_rows(
                min_row=r, max_row=r,
                min_col=start_col, max_col=ec,
                values_only=True,
            )
        )
        return {
            "sheet":     self.sheet.title,
            "row":       r,
            "start_col": start_col,
            "end_col":   ec,
            "values":    list(values),
        }

    def column_values(
        self,
        col: int | str | None = None,
        start_row: int = 1,
        end_row: int | None = None,
    ) -> dict:
        c  = self._nc(col) if col is not None else self.current_col
        er = end_row or self.sheet.max_row or start_row
        self._validate(start_row, c)
        values = [
            row[0]
            for row in self.sheet.iter_rows(
                min_row=start_row, max_row=er,
                min_col=c, max_col=c,
                values_only=True,
            )
        ]
        return {
            "sheet":      self.sheet.title,
            "col":        c,
            "col_letter": _col_letter(c),
            "start_row":  start_row,
            "end_row":    er,
            "values":     values,
        }

    def iter_rows_range(
        self,
        min_row: int = 1,
        max_row: int | None = None,
        min_col: int = 1,
        max_col: int | None = None,
    ) -> list[list]:
        """
        Chunked row iteration — safe for huge sheets.
        Returns at most _CHUNK_ROWS × 2 rows to prevent memory exhaustion.
        """
        mr = max_row or self.sheet.max_row or 1
        mc = max_col or self.sheet.max_column or 1
        cap = min(mr, min_row + _CHUNK_ROWS * 2 - 1)
        return [
            list(row)
            for row in self.sheet.iter_rows(
                min_row=min_row, max_row=cap,
                min_col=min_col, max_col=mc,
                values_only=True,
            )
        ]

    def iter_cols_range(
        self,
        min_col: int = 1,
        max_col: int | None = None,
        min_row: int = 1,
        max_row: int | None = None,
    ) -> list[list]:
        mc = max_col or self.sheet.max_column or 1
        mr = max_row or self.sheet.max_row or 1
        return [
            list(col)
            for col in self.sheet.iter_cols(
                min_col=min_col, max_col=mc,
                min_row=min_row, max_row=mr,
                values_only=True,
            )
        ]

    # ── Bulk helpers for agents ───────────────────────────────────────────────

    def headers(self, row: int = 1, max_col: int | None = None) -> list[str]:
        """Return non-empty header values from the specified row."""
        mc = max_col or self.sheet.max_column or 1
        vals = next(
            self.sheet.iter_rows(
                min_row=row, max_row=row,
                min_col=1, max_col=mc,
                values_only=True,
            )
        )
        return [str(v).strip() for v in vals if _is_meaningful(v)]

    def sample(self, start_row: int = 2, n_rows: int = 8, n_cols: int | None = None) -> list[list]:
        """Return a quick sample from the sheet body."""
        mc = n_cols or min(self.sheet.max_column or 10, 16)
        end = min(self.sheet.max_row or start_row, start_row + n_rows - 1)
        return [
            list(r)
            for r in self.sheet.iter_rows(
                min_row=start_row, max_row=end,
                min_col=1, max_col=mc,
                values_only=True,
            )
        ]

    def close(self) -> None:
        self.workbook.close()

    # ── Private ───────────────────────────────────────────────────────────────

    def _nc(self, col: int | str) -> int:
        if isinstance(col, int):
            return col
        if isinstance(col, str):
            return column_index_from_string(col.upper())
        raise TypeError("Column must be int or letter string.")

    def _validate(self, row: int, col: int) -> None:
        if row < 1:
            raise ValueError("Row index must be >= 1.")
        if col < 1:
            raise ValueError("Column index must be >= 1.")


# ─────────────────────────────────────────────────────────────────────────────
#  Public tool functions
# ─────────────────────────────────────────────────────────────────────────────

def current_time(_: str = "") -> str:
    """Return the current UTC timestamp as an ISO 8601 string."""
    return datetime.utcnow().isoformat() + "Z"


def python_executor(code: str) -> str:
    """Return the received code without executing it (executor disabled)."""
    return f"[python_executor disabled] Received code:\n{code}"


def open_excel(
    file_path: str,
    sheet_name: str | None = None,
    data_only: bool = True,
) -> dict:
    """Open an Excel file and return metadata — fast read_only open."""
    nav = ExcelNavigator(file_path=file_path, sheet_name=sheet_name, data_only=data_only)
    try:
        return nav.workbook_info()
    finally:
        nav.close()


def read_excel_cell(
    file_path: str,
    sheet_name: str,
    row: int,
    col: int | str,
    data_only: bool = True,
) -> dict:
    """Read one cell value."""
    nav = ExcelNavigator(file_path=file_path, sheet_name=sheet_name, data_only=data_only)
    try:
        return nav.cell(row=row, col=col)
    finally:
        nav.close()


def read_excel_row(
    file_path: str,
    sheet_name: str,
    row: int,
    start_col: int = 1,
    end_col: int | None = None,
    data_only: bool = True,
) -> dict:
    """Read one row."""
    nav = ExcelNavigator(file_path=file_path, sheet_name=sheet_name, data_only=data_only)
    try:
        return nav.row_values(row=row, start_col=start_col, end_col=end_col)
    finally:
        nav.close()


def read_excel_column(
    file_path: str,
    sheet_name: str,
    col: int | str,
    start_row: int = 1,
    end_row: int | None = None,
    data_only: bool = True,
) -> dict:
    """Read one column."""
    nav = ExcelNavigator(file_path=file_path, sheet_name=sheet_name, data_only=data_only)
    try:
        return nav.column_values(col=col, start_row=start_row, end_row=end_row)
    finally:
        nav.close()


def read_excel_table(
    file_path: str,
    sheet_name: str,
    min_row: int = 1,
    max_row: int | None = None,
    min_col: int = 1,
    max_col: int | None = None,
    data_only: bool = True,
) -> list[list]:
    """Read a rectangular range, chunked to prevent memory exhaustion."""
    nav = ExcelNavigator(file_path=file_path, sheet_name=sheet_name, data_only=data_only)
    try:
        return nav.iter_rows_range(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
        )
    finally:
        nav.close()


def read_excel_headers(
    file_path: str,
    sheet_name: str,
    row: int = 1,
) -> dict:
    """
    Fast header extraction — opens read_only and reads only row 1.
    Useful for agents that need column names without loading the sheet body.
    """
    nav = ExcelNavigator(file_path=file_path, sheet_name=sheet_name)
    try:
        hdrs = nav.headers(row=row)
        return {"sheet": sheet_name, "header_row": row, "headers": hdrs}
    finally:
        nav.close()


def read_excel_sample(
    file_path: str,
    sheet_name: str,
    start_row: int = 2,
    n_rows: int = 8,
    n_cols: int | None = None,
) -> dict:
    """
    Return a quick sample from the sheet body.
    Designed for agents that need a fast data preview without scanning the whole sheet.
    """
    nav = ExcelNavigator(file_path=file_path, sheet_name=sheet_name)
    try:
        hdrs   = nav.headers(row=1)
        sample = nav.sample(start_row=start_row, n_rows=n_rows, n_cols=n_cols)
        return {
            "sheet":      sheet_name,
            "headers":    hdrs,
            "start_row":  start_row,
            "sample":     sample,
        }
    finally:
        nav.close()


def inspect_workbook(file_path: str) -> dict:
    """
    Profile every sheet in the workbook.

    Strategy:
    1. Open once read_only with data_only=True  → used-area scan, preview, title, header.
    2. Open once read_only with data_only=False → OCR snapshot (formulas + values).
    3. Open once fully                          → visibility check.

    All three opens share the same file — no temp copies, no repeated full loads.
    """
    wb_data    = _open_ro(file_path)
    wb_formula = _open_formulas(file_path)
    wb_full    = _open_full(file_path)

    try:
        profiles: list[dict] = []

        for sn in wb_data.sheetnames:
            ws_data    = wb_data[sn]
            ws_formula = wb_formula[sn]

            vis     = _visibility(wb_full, sn)
            is_hid  = vis in ("hidden", "veryHidden")
            used    = _scan_used_area_fast(ws_data)
            title   = _extract_title_fast(ws_data)
            header  = _detect_header_row_fast(ws_data)
            preview = _preview_fast(ws_data, used)
            ocr     = _extract_ocr_snapshot(ws_data, ws_formula)

            title_norm = _norm(title["title"])
            out_score, src_score, breakdown = _score_sheet(sn, title_norm, used, header, ocr)

            # Hidden sheets get a hard output-score penalty (PROMAT rule)
            if is_hid:
                out_score -= 50
                breakdown.append("hidden sheet (PROMAT penalty −50)")

            profiles.append(
                {
                    "sheet_name":    sn,
                    "is_hidden":     is_hid,
                    "visibility":    vis,
                    "role_guess":    "output" if out_score >= src_score else "source_or_support",
                    "output_score":  out_score,
                    "source_score":  src_score,
                    "score_breakdown": breakdown,
                    "used":          used,
                    "title":         title,
                    "header":        header,
                    "preview":       preview,
                    "ocr":           ocr,
                }
            )

        output_candidates = sorted(profiles, key=lambda x: x["output_score"], reverse=True)
        source_candidates = sorted(profiles, key=lambda x: x["source_score"], reverse=True)

        main_output = output_candidates[0]["sheet_name"] if output_candidates else None
        main_source = source_candidates[0]["sheet_name"] if source_candidates else None

        return {
            "active_sheet":       wb_data.active.title if wb_data.active else None,
            "profiles":           profiles,
            "main_output_sheet":  main_output,
            "main_source_sheet":  main_source,
            "output_candidates":  [
                {"sheet": p["sheet_name"], "score": p["output_score"], "title": p["title"]["title"]}
                for p in output_candidates[:10]
            ],
            "source_candidates":  [
                {"sheet": p["sheet_name"], "score": p["source_score"], "title": p["title"]["title"]}
                for p in source_candidates[:10]
            ],
        }
    finally:
        wb_data.close()
        wb_formula.close()
        wb_full.close()


def detect_main_sheet(file_path: str) -> dict:
    """
    Deterministic main-sheet detector.
    Wraps inspect_workbook and maps result to the canonical PROMAT output shape.
    """
    ins = inspect_workbook(file_path)
    return {
        "main_sheet":          ins["main_output_sheet"],
        "main_output_sheet":   ins["main_output_sheet"],
        "main_source_sheet":   ins["main_source_sheet"],
        "active_sheet":        ins["active_sheet"],
        "output_candidates":   ins["output_candidates"],
        "source_candidates":   ins["source_candidates"],
        "profiles":            ins["profiles"],
    }


def sheet_ocr_snapshot(
    file_path: str,
    sheet_name: str,
) -> dict:
    """
    Standalone OCR snapshot for a single sheet.
    Agents can call this directly when they need deep formula + value extraction
    for one specific sheet without re-profiling the whole workbook.
    """
    wb_data    = _open_ro(file_path)
    wb_formula = _open_formulas(file_path)
    wb_full    = _open_full(file_path)

    try:
        ws_data    = wb_data[sheet_name]
        ws_formula = wb_formula[sheet_name]

        vis    = _visibility(wb_full, sheet_name)
        used   = _scan_used_area_fast(ws_data)
        title  = _extract_title_fast(ws_data)
        header = _detect_header_row_fast(ws_data)
        ocr    = _extract_ocr_snapshot(ws_data, ws_formula)

        return {
            "sheet_name":  sheet_name,
            "is_hidden":   vis in ("hidden", "veryHidden"),
            "visibility":  vis,
            "used":        used,
            "title":       title,
            "header":      header,
            **ocr,
        }
    finally:
        wb_data.close()
        wb_formula.close()
        wb_full.close()


def list_sheets_with_visibility(file_path: str) -> list[dict]:
    """
    Return all sheet names with their visibility status.
    Uses a full open (required to read sheet visibility).
    Agents should call this first before any other operation.
    """
    wb = _open_full(file_path)
    try:
        return [
            {
                "name":       sn,
                "visibility": _visibility(wb, sn),
                "is_hidden":  _visibility(wb, sn) in ("hidden", "veryHidden"),
            }
            for sn in wb.sheetnames
        ]
    finally:
        wb.close()